import pandas as pd
import io
import zipfile
import streamlit as st
import xml.etree.ElementTree as ET
import re
import os
from datetime import datetime
import openpyxl
from copy import copy

# --- IMPORTAÇÃO DOS MÓDULOS ESPECIALISTAS (MANTIDOS ÍNTEGROS) ---
try:
    from audit_resumo import gerar_aba_resumo
    from Auditorias.audit_icms import processar_icms
    from Auditorias.audit_ipi import processar_ipi
    from Auditorias.audit_pis_cofins import processar_pc
    from Auditorias.audit_difal import processar_difal
    from Apuracoes.apuracao_difal import gerar_resumo_uf
    from Gerenciais.audit_gerencial import gerar_abas_gerenciais
except ImportError as e:
    st.error(f"⚠️ Erro Crítico de Dependência: Verifique se as pastas Auditorias, Apuracoes e Gerenciais estão no diretório. Detalhe: {e}")

# --- UTILITÁRIOS DE TRATAMENTO DE DADOS E CONVERSÃO ---
def safe_float(v):
    """Converte valores para float de forma segura, tratando R$, vírgulas e strings nulas."""
    if v is None or pd.isna(v): return 0.0
    txt = str(v).strip().upper()
    if txt in ['NT', '', 'N/A', 'ISENTO', 'NULL', 'ZERO', '-', ' ']: return 0.0
    try:
        txt = txt.replace('R$', '').replace(' ', '').replace('%', '').strip()
        if ',' in txt and '.' in txt: txt = txt.replace('.', '').replace(',', '.')
        elif ',' in txt: txt = txt.replace(',', '.')
        return round(float(txt), 4)
    except:
        return 0.0

def buscar_tag_recursiva(tag_alvo, no):
    """Busca uma tag em qualquer nível do XML sem depender de namespaces fixos."""
    if no is None: return ""
    for elemento in no.iter():
        tag_nome = elemento.tag.split('}')[-1]
        if tag_nome == tag_alvo:
            return elemento.text if elemento.text else ""
    return ""

def tratar_ncm_texto(ncm):
    """Normaliza o NCM como texto puro, preservando zeros à esquerda e removendo pontuação."""
    if pd.isna(ncm) or ncm == "": return ""
    return re.sub(r'\D', '', str(ncm)).strip()

# --- MOTOR DE PROCESSAMENTO XML (ESTRUTURA DEEP-PROCESS) ---
def processar_conteudo_xml(content, dados_lista, cnpj_empresa_auditada):
    """Realiza a leitura completa de cada item do XML, mapeando todas as tags tributárias."""
    try:
        xml_str = content.decode('utf-8', errors='replace')
        xml_str = re.sub(r'\sxmlns(:\w+)?="[^"]+"', '', xml_str) 
        root = ET.fromstring(xml_str)
        
        inf = root.find('.//infNFe')
        if inf is None: return 
        
        ide = root.find('.//ide')
        emit = root.find('.//emit')
        dest = root.find('.//dest')
        
        cnpj_emit = re.sub(r'\D', '', buscar_tag_recursiva('CNPJ', emit))
        cnpj_alvo = re.sub(r'\D', '', str(cnpj_empresa_auditada))
        tipo_nf = buscar_tag_recursiva('tpNF', ide)
        
        tipo_operacao = "SAIDA" if (cnpj_emit == cnpj_alvo and tipo_nf == '1') else "ENTRADA"
        
        chave = inf.attrib.get('Id', '')[3:]
        n_nf = buscar_tag_recursiva('nNF', ide)
        dt_emi = buscar_tag_recursiva('dhEmi', ide) or buscar_tag_recursiva('dEmi', ide)

        for det in root.findall('.//det'):
            prod = det.find('prod'); imp = det.find('imposto')
            if prod is None or imp is None: continue
            
            icms_no = det.find('.//ICMS'); ipi_no = det.find('.//IPI')
            pis_no = det.find('.//PIS'); cof_no = det.find('.//COFINS')

            linha = {
                "TIPO_SISTEMA": tipo_operacao,
                "CHAVE_ACESSO": str(chave).strip(),
                "NUM_NF": n_nf,
                "DATA_EMISSAO": dt_emi,
                "CNPJ_EMIT": cnpj_emit,
                "CNPJ_DEST": re.sub(r'\D', '', buscar_tag_recursiva('CNPJ', dest)),
                "UF_EMIT": buscar_tag_recursiva('UF', emit),
                "UF_DEST": buscar_tag_recursiva('UF', dest),
                "CFOP": buscar_tag_recursiva('CFOP', prod),
                "NCM": tratar_ncm_texto(buscar_tag_recursiva('NCM', prod)),
                "VPROD": safe_float(buscar_tag_recursiva('vProd', prod)),
                "ORIGEM": buscar_tag_recursiva('orig', icms_no),
                "CST-ICMS": buscar_tag_recursiva('CST', icms_no) or buscar_tag_recursiva('CSOSN', icms_no),
                "BC-ICMS": safe_float(buscar_tag_recursiva('vBC', icms_no)),
                "ALQ-ICMS": safe_float(buscar_tag_recursiva('pICMS', icms_no)),
                "VLR-ICMS": safe_float(buscar_tag_recursiva('vICMS', icms_no)),
                "BC-ICMS-ST": safe_float(buscar_tag_recursiva('vBCST', icms_no)),
                "VAL-ICMS-ST": safe_float(buscar_tag_recursiva('vICMSST', icms_no)),
                "IE_SUBST": str(buscar_tag_recursiva('IEST', icms_no)).strip(),
                "ALQ-IPI": safe_float(buscar_tag_recursiva('pIPI', ipi_no)),
                "VLR-IPI": safe_float(buscar_tag_recursiva('vIPI', ipi_no)),
                "CST-IPI": buscar_tag_recursiva('CST', ipi_no),
                "VAL-IBS": safe_float(buscar_tag_recursiva('vIBS', imp)),
                "VAL-CBS": safe_float(buscar_tag_recursiva('vCBS', imp)),
                "CST-PIS": buscar_tag_recursiva('CST', pis_no),
                "VLR-PIS": safe_float(buscar_tag_recursiva('vPIS', pis_no)),
                "CST-COFINS": buscar_tag_recursiva('CST', cof_no),
                "VLR-COFINS": safe_float(buscar_tag_recursiva('vCOFINS', cof_no)),
                "VAL-FCP": safe_float(buscar_tag_recursiva('vFCP', imp)),
                "VAL-DIFAL": safe_float(buscar_tag_recursiva('vICMSUFDest', imp)) + safe_float(buscar_tag_recursiva('vFCPUFDest', imp)),
                "VAL-FCP-DEST": safe_float(buscar_tag_recursiva('vFCPUFDest', imp))
            }
            dados_lista.append(linha)
    except Exception as e:
        print(f"Erro no processamento de item XML: {e}")

# --- EXTRAÇÃO RECURSIVA DE ZIP ---
def extrair_dados_xml_recursivo(files, cnpj_empresa_auditada):
    dados_lista = []
    if not files: return pd.DataFrame(), pd.DataFrame()
    lista_trabalho = files if isinstance(files, list) else [files]
    
    def ler_zip(zip_data):
        try:
            with zipfile.ZipFile(zip_data) as z:
                for filename in z.namelist():
                    if filename.lower().endswith('.xml'):
                        with z.open(filename) as f:
                            processar_conteudo_xml(f.read(), dados_lista, cnpj_empresa_auditada)
                    elif filename.lower().endswith('.zip'):
                        ler_zip(io.BytesIO(z.read(filename)))
        except Exception as e:
            st.error(f"Erro ao abrir arquivo ZIP: {e}")

    for f in lista_trabalho:
        f.seek(0)
        if f.name.lower().endswith('.xml'):
            processar_conteudo_xml(f.read(), dados_lista, cnpj_empresa_auditada)
        elif f.name.lower().endswith('.zip'):
            ler_zip(f)
            
    df_total = pd.DataFrame(dados_lista)
    if df_total.empty: return pd.DataFrame(), pd.DataFrame()
    return df_total[df_total['TIPO_SISTEMA'] == "ENTRADA"].copy(), df_total[df_total['TIPO_SISTEMA'] == "SAIDA"].copy()

# --- GERAÇÃO DO EXCEL FINAL (ESTRUTURA COMPLETA) ---
def gerar_excel_final(df_xe, df_xs, ae, as_f, ge, gs, cod_cliente, regime, is_ret):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # 1. Aba Resumo
        try: gerar_aba_resumo(writer)
        except Exception as e: st.warning(f"Aviso ao gerar aba resumo: {e}")
        
        # 2. Chamada das Gerenciais (Módulo Externo)
        try: gerar_abas_gerenciais(writer, ge, gs)
        except Exception as e: st.error(f"Erro no módulo de gerenciais: {e}")

        # 3. Processamento de Auditorias XML
        if not df_xs.empty:
            st_map = {}
            for f_auth in ([ae] if ae else []) + ([as_f] if as_f else []):
                if not f_auth: continue
                try:
                    f_auth.seek(0)
                    if f_auth.name.endswith('.xlsx'): df_a = pd.read_excel(f_auth, header=None)
                    else: df_a = pd.read_csv(f_auth, header=None, sep=None, engine='python')
                    
                    df_a[0] = df_a[0].astype(str).str.replace('NFe', '').str.strip()
                    st_map.update(df_a.set_index(0)[5].to_dict())
                except: continue
            
            df_xs['Situação Nota'] = df_xs['CHAVE_ACESSO'].map(st_map).fillna('⚠️ N/Encontrada')
            
            processar_icms(df_xs, writer, cod_cliente, df_xe)
            processar_ipi(df_xs, writer, cod_cliente)
            processar_pc(df_xs, writer, cod_cliente, regime)
            processar_difal(df_xs, writer)
            try: gerar_resumo_uf(df_xs, writer, df_xe)
            except: pass
            
    # Lógica de clonagem para RET MG
    if is_ret:
        try:
            caminho_modelo = f"RET/{cod_cliente}-RET_MG.xlsx"
            if os.path.exists(caminho_modelo):
                output.seek(0)
                wb_final = openpyxl.load_workbook(output)
                wb_modelo = openpyxl.load_workbook(caminho_modelo, data_only=False)
                for sheet_name in wb_modelo.sheetnames:
                    if sheet_name not in wb_final.sheetnames:
                        source = wb_modelo[sheet_name]; target = wb_final.create_sheet(sheet_name)
                        for row in source.iter_rows():
                            for cell in row:
                                new_cell = target.cell(row=cell.row, column=cell.column, value=cell.value)
                                if cell.has_style:
                                    new_cell.font, new_cell.border, new_cell.fill, new_cell.number_format, new_cell.alignment = copy(cell.font), copy(cell.border), copy(cell.fill), copy(cell.number_format), copy(cell.alignment)
                output_ret = io.BytesIO()
                wb_final.save(output_ret)
                return output_ret.getvalue()
        except Exception as e:
            st.error(f"Erro na integração RET MG: {e}")

    return output.getvalue()
