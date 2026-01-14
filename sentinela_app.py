import pandas as pd
import io
import zipfile
import streamlit as st
import xml.etree.ElementTree as ET
import re
import os
import requests
from datetime import datetime
import openpyxl
from copy import copy

# --- IMPORTAÇÃO DOS MÓDULOS ESPECIALISTAS ---
try:
    from audit_resumo import gerar_aba_resumo
    from Auditorias.audit_icms import processar_icms
    from Auditorias.audit_ipi import processar_ipi
    from Auditorias.audit_pis_cofins import processar_pc
    from Auditorias.audit_difal import processar_difal
    try: from Apuracoes.apuracao_difal import gerar_resumo_uf
    except ImportError: from Apuracoes.apuracao_difal import gerar_resumo_uf
except ImportError as e:
    st.error(f"Erro Crítico de Dependência: {e}")

def safe_float(v):
    if v is None or pd.isna(v): return 0.0
    txt = str(v).strip().upper()
    try:
        txt = txt.replace('R$', '').replace(' ', '').replace('%', '').strip()
        if ',' in txt and '.' in txt: txt = txt.replace('.', '').replace(',', '.')
        elif ',' in txt: txt = txt.replace(',', '.')
        return round(float(txt), 4)
    except: return 0.0

def buscar_tag_recursiva(tag_alvo, no):
    if no is None: return ""
    for elemento in no.iter():
        if elemento.tag.split('}')[-1] == tag_alvo:
            return elemento.text if elemento.text else ""
    return ""

def processar_conteudo_xml(content, dados_lista, cnpj_empresa_auditada):
    try:
        xml_str = content.decode('utf-8', errors='replace')
        xml_str = re.sub(r'\sxmlns(:\w+)?="[^"]+"', '', xml_str)
        root = ET.fromstring(xml_str)
        inf = root.find('.//infNFe')
        if inf is None: return 
        ide = root.find('.//ide'); emit = root.find('.//emit')
        tp_nf = buscar_tag_recursiva('tpNF', ide)
        cnpj_emit = re.sub(r'\D', '', buscar_tag_recursiva('CNPJ', emit))
        cnpj_alvo = re.sub(r'\D', '', str(cnpj_empresa_auditada))
        tipo_operacao = "SAIDA" if (cnpj_emit == cnpj_alvo and tp_nf == '1') else "ENTRADA"
        chave = inf.attrib.get('Id', '')[3:]; n_nf = buscar_tag_recursiva('nNF', ide)
        for det in root.findall('.//det'):
            prod = det.find('prod'); imp = det.find('imposto')
            if prod is None or imp is None: continue
            linha = {
                "TIPO_SISTEMA": tipo_operacao, "CHAVE_ACESSO": str(chave).strip(), "NUM_NF": n_nf,
                "CNPJ_EMIT": buscar_tag_recursiva('CNPJ', emit), "CNPJ_DEST": buscar_tag_recursiva('CNPJ', root.find('.//dest')),
                "UF_EMIT": buscar_tag_recursiva('UF', emit), "UF_DEST": buscar_tag_recursiva('UF', root.find('.//dest')),
                "CFOP": buscar_tag_recursiva('CFOP', prod), "NCM": re.sub(r'\D', '', buscar_tag_recursiva('NCM', prod)).zfill(8),
                "VPROD": safe_float(buscar_tag_recursiva('vProd', prod)), "CST-ICMS": buscar_tag_recursiva('CST', det.find('.//ICMS')) or buscar_tag_recursiva('CSOSN', det.find('.//ICMS')),
                "VAL-ICMS-ST": safe_float(buscar_tag_recursiva('vICMSST', imp)), "BC-ICMS-ST": safe_float(buscar_tag_recursiva('vBCST', imp)),
                "IE_SUBST": buscar_tag_recursiva('IEST', root) or buscar_tag_recursiva('IE_SUBST', root)
            }
            dados_lista.append(linha)
    except: pass

def extrair_dados_xml_recursivo(files, cnpj_empresa_auditada):
    dados_lista = []
    lista_trabalho = files if isinstance(files, list) else [files]
    def ler_recursivo(conteudo):
        try:
            with zipfile.ZipFile(io.BytesIO(conteudo)) as z:
                for filename in z.namelist():
                    if filename.lower().endswith('.xml'):
                        with z.open(filename) as f: processar_conteudo_xml(f.read(), dados_lista, cnpj_empresa_auditada)
                    elif filename.lower().endswith('.zip'):
                        with z.open(filename) as f: ler_recursivo(f.read())
        except: pass
    for f in lista_trabalho:
        content = f.read()
        if f.name.lower().endswith('.xml'): processar_conteudo_xml(content, dados_lista, cnpj_empresa_auditada)
        elif f.name.lower().endswith('.zip'): ler_recursivo(content)
    df_total = pd.DataFrame(dados_lista)
    if df_total.empty: return pd.DataFrame(), pd.DataFrame()
    return df_total[df_total['TIPO_SISTEMA'] == "ENTRADA"].copy(), df_total[df_total['TIPO_SISTEMA'] == "SAIDA"].copy()

def baixar_arquivo_github(caminho_relativo):
    token = st.secrets.get("GITHUB_TOKEN")
    repo = st.secrets.get("GITHUB_REPO")
    url = f"https://raw.githubusercontent.com/{repo}/main/{caminho_relativo}"
    headers = {"Authorization": f"token {token}"}
    try:
        res = requests.get(url, headers=headers)
        if res.status_code == 200: return io.BytesIO(res.content)
    except: pass
    return None

def gerar_excel_final(df_xe, df_xs, ae, as_f, ge, gs, cod_cliente, regime, is_ret):
    output = io.BytesIO()
    cols_ent = ["NUM_NF","DATA_EMISSAO","CNPJ","UF","VLR_NF","AC","CFOP","COD_PROD","DESCR","NCM","UNID","VUNIT","QTDE","VPROD","DESC","FRETE","SEG","DESP","VC","CST-ICMS","BC-ICMS","VLR-ICMS","BC-ICMS-ST","ICMS-ST","VLR_IPI","CST_PIS","BC_PIS","VLR_PIS","CST_COF","BC_COF","VLR_COF"]
    cols_sai = ["NF","DATA_EMISSAO","CNPJ","Ufp","VC","AC","CFOP","COD_ITEM","DESC_ITEM","NCM","UND","VUNIT","QTDE","VITEM","DESC","FRETE","SEG","OUTRAS","VC_ITEM","CST","BC_ICMS","ALIQ_ICMS","ICMS","BC_ICMSST","ICMSST","IPI","CST_PIS","BC_PIS","PIS","CST_COF","BC_COF","COF"]
    
    def ler_csv(arquivos, colunas):
        if not arquivos: return pd.DataFrame()
        dfs = []
        for f in (arquivos if isinstance(arquivos, list) else [arquivos]):
            f.seek(0)
            try:
                raw = f.read().decode('latin1', errors='replace')
                data = [re.split(r'\t|;', l) for l in raw.splitlines() if l.strip()]
                df = pd.DataFrame(data, columns=colunas)
                for c in df.columns:
                    if any(k in c.upper() for k in ['VLR', 'BC', 'VAL', 'VC', 'QTDE', 'VUNIT', 'ICMS']): df[c] = df[c].apply(safe_float)
                dfs.append(df)
            except: pass
        return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

    df_ge = ler_csv(ge, cols_ent); df_gs = ler_csv(gs, cols_sai)
    
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book
        try: gerar_aba_resumo(writer)
        except: pass
        
        def gravar_tabela(df, sheet, name):
            if df.empty: return
            ws = workbook.add_worksheet(sheet); writer.sheets[sheet] = ws
            ws.add_table(0, 0, len(df), len(df.columns)-1, {'data': df.values.tolist(), 'columns': [{'header': c} for c in df.columns], 'name': name, 'style': 'TableStyleMedium2'})
        
        gravar_tabela(df_ge, 'GERENCIAL_ENTRADAS', 'TabEnt'); gravar_tabela(df_gs, 'GERENCIAL_SAIDAS', 'TabSai')
        
        if not df_xs.empty:
            st_map = {}
            for f in ( (ae if ae else []) + (as_f if as_f else []) ):
                try:
                    f.seek(0); df_a = pd.read_excel(f, header=None) if f.name.endswith('.xlsx') else pd.read_csv(f, header=None, sep=None, engine='python')
                    df_a[0] = df_a[0].astype(str).str.replace('NFe', '').str.strip()
                    st_map.update(df_a.set_index(0)[5].to_dict())
                except: pass
            df_xs['Situação Nota'] = df_xs['CHAVE_ACESSO'].map(st_map).fillna('⚠️ N/Encontrada')
            
            processar_icms(df_xs, writer, cod_cliente)
            processar_ipi(df_xs, writer, cod_cliente)
            processar_pc(df_xs, writer, cod_cliente, regime)
            processar_difal(df_xs, writer)
            try: gerar_resumo_uf(df_xs, writer, df_xe)
            except: pass

    if is_ret:
        ret_file = baixar_arquivo_github(f"RET/{cod_cliente}-RET_MG.xlsx")
        if ret_file:
            try:
                output.seek(0); wb_f = openpyxl.load_workbook(output); wb_m = openpyxl.load_workbook(ret_file, data_only=False)
                for sn in wb_m.sheetnames:
                    if sn not in ['GERENCIAL_ENTRADAS', 'GERENCIAL_SAIDAS']:
                        source = wb_m[sn]; target = wb_f.create_sheet(sn)
                        for r in source.iter_rows():
                            for c in r:
                                nc = target.cell(row=c.row, column=c.column, value=c.value)
                                if c.has_style: nc.font, nc.border, nc.fill, nc.number_format, nc.alignment = copy(c.font), copy(c.border), copy(c.fill), copy(c.number_format), copy(c.alignment)
                        for col, dim in source.column_dimensions.items(): target.column_dimensions[col].width = dim.width
                out_r = io.BytesIO(); wb_f.save(out_r); return out_r.getvalue()
            except: pass
            
    return output.getvalue()
